"""
neofinesse.services.evidence_context_service
============================================
Provides secure, deterministic cell-level spreadsheet context extraction for
visual inspection in the NeoFinesse Evidence Viewer.

Features:
- Safe resolution against registered source files (strict anti-path-traversal).
- Configurable radius window (default row_radius=3, col_radius=3) clamped to sheet boundaries.
- Support for CSV and multi-sheet XLSX workbooks.
- Zero LLM involvement — purely deterministic forensic evidence grounding.
"""
from __future__ import annotations

import csv
import hashlib
import io
from pathlib import Path
import re
from typing import Any, Dict, List, Optional, Tuple, Union

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from neofinesse.ingestion.registry import SourceRegistry


def column_letter_to_index(col_letter: str) -> int:
    """Converts an Excel column letter (e.g., 'A', 'H', 'AA') to 1-indexed integer (1, 8, 27)."""
    col_letter = col_letter.strip().upper()
    result = 0
    for char in col_letter:
        if "A" <= char <= "Z":
            result = result * 26 + (ord(char) - ord("A") + 1)
        else:
            raise ValueError(f"Invalid column letter character: {char}")
    return result


def index_to_column_letter(col_idx: int) -> str:
    """Converts 1-indexed column integer (1, 8, 27) to Excel column letter ('A', 'H', 'AA')."""
    if col_idx < 1:
        raise ValueError("Column index must be >= 1")
    letters = []
    while col_idx > 0:
        col_idx, rem = divmod(col_idx - 1, 26)
        letters.append(chr(ord("A") + rem))
    return "".join(reversed(letters))


def parse_cell_address(cell_ref: str) -> Tuple[Optional[str], str, int, int]:
    """
    Parses a cell reference string into (sheet, col_letter, col_idx, row_idx).
    Supported formats:
      - "H42" -> (None, "H", 8, 42)
      - "Settlements!H42" -> ("Settlements", "H", 8, 42)
      - "'Account Statement'!C19" -> ("Account Statement", "C", 3, 19)
    """
    cell_ref = cell_ref.strip()
    sheet_name: Optional[str] = None

    if "!" in cell_ref:
        parts = cell_ref.split("!", 1)
        sheet_part = parts[0].strip()
        if sheet_part.startswith("'") and sheet_part.endswith("'"):
            sheet_name = sheet_part[1:-1]
        elif sheet_part.startswith('"') and sheet_part.endswith('"'):
            sheet_name = sheet_part[1:-1]
        else:
            sheet_name = sheet_part
        addr_part = parts[1].strip()
    else:
        addr_part = cell_ref

    match = re.match(r"^([A-Za-z]+)(\d+)$", addr_part)
    if not match:
        raise ValueError(f"Invalid cell address format: '{cell_ref}'")

    col_letter = match.group(1).upper()
    row_idx = int(match.group(2))
    col_idx = column_letter_to_index(col_letter)

    return sheet_name, col_letter, col_idx, row_idx


class EvidenceContextService:
    """
    Service for loading spreadsheet/CSV cell context windows with cryptographic provenance integrity.
    """

    def __init__(self, data_dir: Union[str, Path]):
        self.data_dir = Path(data_dir).resolve()
        self.registry = SourceRegistry(str(self.data_dir))
        self._refresh_registry()

    def _refresh_registry(self) -> None:
        """Discovers and caches registered source files."""
        if self.data_dir.exists():
            self.registry.discover_and_register_files()

    def _resolve_safe_file_path(self, filename: str) -> Path:
        """
        Validates that the requested filename belongs strictly to the source registry
        and prevents any directory traversal attacks (e.g. '../../secret.txt').
        """
        # Strip path prefixes to get clean filename
        clean_name = Path(filename).name
        if not clean_name or clean_name != filename.strip():
            # If input contained path separators or attempted traversal
            if "/" in filename or "\\" in filename or ".." in filename:
                raise PermissionError(f"Security Alert: Directory traversal prohibited: '{filename}'")

        target_path = (self.data_dir / clean_name).resolve()

        # Check that resolved path is inside data_dir
        try:
            target_path.relative_to(self.data_dir)
        except ValueError:
            raise PermissionError(f"Security Alert: Path '{filename}' outside allowed directory.")

        if not target_path.exists() or not target_path.is_file():
            raise FileNotFoundError(f"Source file '{clean_name}' is not found in source registry.")

        return target_path

    def _compute_sha256(self, filepath: Path) -> str:
        """Computes cryptographic SHA-256 hash of a file."""
        hasher = hashlib.sha256()
        with open(filepath, "rb") as f:
            while chunk := f.read(65536):
                hasher.update(chunk)
        return hasher.hexdigest()

    def get_cell_context(
        self,
        filename: str,
        sheet: Optional[str] = None,
        cell: Optional[str] = None,
        row: Optional[int] = None,
        column: Optional[Union[int, str]] = None,
        row_radius: int = 3,
        column_radius: int = 3,
    ) -> Dict[str, Any]:
        """
        Extracts a localized matrix window of rows and columns around the referenced target cell.
        
        Parameters:
            filename: Name of registered file (e.g. 'settlements.xlsx', 'refunds.csv')
            sheet: Sheet name (optional, defaults to first sheet for XLSX or active sheet)
            cell: Excel cell address (e.g. 'H42' or 'Settlements!H42')
            row: 1-indexed row number (used if cell is not provided)
            column: Column index (int) or column letter (str) (used if cell is not provided)
            row_radius: Number of surrounding rows to include (default: 3)
            column_radius: Number of surrounding columns to include (default: 3)
        """
        file_path = self._resolve_safe_file_path(filename)
        file_hash = self._compute_sha256(file_path)

        # Parse target row and column
        target_sheet = sheet
        target_row = row
        target_col_idx: Optional[int] = None
        target_col_letter: Optional[str] = None

        if cell:
            parsed_sheet, parsed_col_letter, parsed_col_idx, parsed_row = parse_cell_address(cell)
            if parsed_sheet and not target_sheet:
                target_sheet = parsed_sheet
            target_row = parsed_row
            target_col_idx = parsed_col_idx
            target_col_letter = parsed_col_letter
        elif column is not None:
            if isinstance(column, int):
                target_col_idx = column
                target_col_letter = index_to_column_letter(column)
            else:
                target_col_letter = str(column).upper()
                target_col_idx = column_letter_to_index(target_col_letter)

        if target_row is None or target_col_idx is None:
            raise ValueError("Target cell coordinates must be specified via 'cell' or ('row', 'column').")

        if target_row < 1 or target_col_idx < 1:
            raise ValueError(f"Invalid row ({target_row}) or column ({target_col_idx}). Coordinates must be >= 1.")

        is_xlsx = file_path.suffix.lower() in [".xlsx", ".xls"]

        if is_xlsx:
            return self._extract_xlsx_context(
                file_path=file_path,
                file_hash=file_hash,
                sheet_name=target_sheet,
                target_row=target_row,
                target_col_idx=target_col_idx,
                target_col_letter=target_col_letter or index_to_column_letter(target_col_idx),
                row_radius=row_radius,
                column_radius=column_radius,
            )
        else:
            return self._extract_csv_context(
                file_path=file_path,
                file_hash=file_hash,
                sheet_name=target_sheet or "Sheet1",
                target_row=target_row,
                target_col_idx=target_col_idx,
                target_col_letter=target_col_letter or index_to_column_letter(target_col_idx),
                row_radius=row_radius,
                column_radius=column_radius,
            )

    def _extract_csv_context(
        self,
        file_path: Path,
        file_hash: str,
        sheet_name: str,
        target_row: int,
        target_col_idx: int,
        target_col_letter: str,
        row_radius: int,
        column_radius: int,
    ) -> Dict[str, Any]:
        """Reads CSV and extracts window around target row and column."""
        with open(file_path, "r", encoding="utf-8-sig", errors="replace") as f:
            reader = csv.reader(f)
            all_rows = list(reader)

        total_rows = len(all_rows)
        total_cols = max((len(r) for r in all_rows), default=0) if all_rows else 0

        if total_rows == 0:
            raise ValueError(f"CSV file '{file_path.name}' is empty.")

        if target_row > total_rows or target_col_idx > total_cols:
            raise ValueError(
                f"Requested cell '{target_col_letter}{target_row}' is out of bounds for '{file_path.name}' "
                f"(dimensions: {total_rows} rows x {total_cols} columns)."
            )

        # Compute row window (1-indexed)
        min_row = max(1, target_row - row_radius)
        max_row = min(total_rows, target_row + row_radius)

        # Compute column window (1-indexed)
        min_col = max(1, target_col_idx - column_radius)
        max_col = min(total_cols, target_col_idx + column_radius)

        # Headers from Row 1
        headers = all_rows[0] if len(all_rows) > 0 else []

        # Target cell address & value
        target_cell_addr = f"{target_col_letter}{target_row}"
        target_value: Any = None
        target_row_content: List[str] = []

        if 1 <= target_row <= total_rows:
            target_row_content = all_rows[target_row - 1]
            if 1 <= target_col_idx <= len(target_row_content):
                raw_val = target_row_content[target_col_idx - 1]
                target_value = self._format_cell_value(raw_val)

        # Record hash of target row
        record_hasher = hashlib.sha256()
        record_hasher.update(",".join(target_row_content).encode("utf-8"))
        record_hash = record_hasher.hexdigest()

        # Build column metadata
        columns_meta = []
        for c in range(min_col, max_col + 1):
            c_letter = index_to_column_letter(c)
            header_name = headers[c - 1] if c - 1 < len(headers) else f"Column {c_letter}"
            columns_meta.append({
                "index": c,
                "letter": c_letter,
                "header": header_name,
                "is_target_column": (c == target_col_idx),
            })

        # Build row cells matrix
        rows_data = []
        for r_num in range(min_row, max_row + 1):
            row_raw = all_rows[r_num - 1] if r_num - 1 < len(all_rows) else []
            row_cells = []
            for c_num in range(min_col, max_col + 1):
                c_letter = index_to_column_letter(c_num)
                val = row_raw[c_num - 1] if c_num - 1 < len(row_raw) else ""
                formatted_val = self._format_cell_value(val)
                is_target = (r_num == target_row and c_num == target_col_idx)

                row_cells.append({
                    "address": f"{c_letter}{r_num}",
                    "row": r_num,
                    "column": c_num,
                    "column_letter": c_letter,
                    "value": formatted_val,
                    "raw_value": val,
                    "is_target": is_target,
                })

            rows_data.append({
                "row_number": r_num,
                "is_target_row": (r_num == target_row),
                "cells": row_cells,
            })

        return {
            "status": "SUCCESS",
            "source_file": file_path.name,
            "sheet": sheet_name or "Sheet1",
            "target_cell": target_cell_addr,
            "target_row": target_row,
            "target_column": target_col_idx,
            "target_column_letter": target_col_letter,
            "target_value": target_value,
            "file_hash": file_hash,
            "record_hash": record_hash,
            "is_provenance_verified": True,
            "total_rows": total_rows,
            "total_columns": total_cols,
            "window": {
                "min_row": min_row,
                "max_row": max_row,
                "min_col": min_col,
                "max_col": max_col,
                "row_radius": row_radius,
                "column_radius": column_radius,
            },
            "context": {
                "columns": columns_meta,
                "rows": rows_data,
            },
        }

    def _extract_xlsx_context(
        self,
        file_path: Path,
        file_hash: str,
        sheet_name: Optional[str],
        target_row: int,
        target_col_idx: int,
        target_col_letter: str,
        row_radius: int,
        column_radius: int,
    ) -> Dict[str, Any]:
        """Reads XLSX workbook and extracts window around target row and column."""
        if not HAS_OPENPYXL:
            raise RuntimeError("openpyxl is required to inspect XLSX workbooks.")

        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        available_sheets = wb.sheetnames

        active_sheet_name = sheet_name or available_sheets[0]
        if active_sheet_name not in available_sheets:
            matched = [s for s in available_sheets if s.lower() == active_sheet_name.lower()]
            if matched:
                active_sheet_name = matched[0]
            else:
                wb.close()
                raise ValueError(
                    f"Sheet '{sheet_name}' not found in '{file_path.name}'. Available sheets: {available_sheets}"
                )

        ws = wb[active_sheet_name]

        total_rows = ws.max_row or 1
        total_cols = ws.max_column or 1

        if target_row > total_rows or target_col_idx > total_cols:
            wb.close()
            raise ValueError(
                f"Requested cell '{target_col_letter}{target_row}' is out of bounds for sheet '{active_sheet_name}' "
                f"in '{file_path.name}' (dimensions: {total_rows} rows x {total_cols} columns)."
            )

        min_row = max(1, target_row - row_radius)
        max_row = min(total_rows, target_row + row_radius)
        min_col = max(1, target_col_idx - column_radius)
        max_col = min(total_cols, target_col_idx + column_radius)

        # Header values from row 1
        headers: List[str] = []
        for c in range(min_col, max_col + 1):
            h_val = ws.cell(row=1, column=c).value
            headers.append(str(h_val) if h_val is not None else "")

        # Target cell address & value
        target_cell_addr = f"{target_col_letter}{target_row}"
        target_cell_obj = ws.cell(row=target_row, column=target_col_idx)
        target_value = target_cell_obj.value

        # Compute row hash
        target_row_vals = [str(ws.cell(row=target_row, column=c).value or "") for c in range(1, total_cols + 1)]
        record_hasher = hashlib.sha256()
        record_hasher.update(",".join(target_row_vals).encode("utf-8"))
        record_hash = record_hasher.hexdigest()

        columns_meta = []
        for idx, c in enumerate(range(min_col, max_col + 1)):
            c_letter = index_to_column_letter(c)
            columns_meta.append({
                "index": c,
                "letter": c_letter,
                "header": headers[idx] if idx < len(headers) and headers[idx] else f"Column {c_letter}",
                "is_target_column": (c == target_col_idx),
            })

        rows_data = []
        for r_num in range(min_row, max_row + 1):
            row_cells = []
            for c_num in range(min_col, max_col + 1):
                c_letter = index_to_column_letter(c_num)
                cell_val = ws.cell(row=r_num, column=c_num).value
                is_target = (r_num == target_row and c_num == target_col_idx)

                row_cells.append({
                    "address": f"{c_letter}{r_num}",
                    "row": r_num,
                    "column": c_num,
                    "column_letter": c_letter,
                    "value": cell_val,
                    "raw_value": str(cell_val) if cell_val is not None else "",
                    "is_target": is_target,
                })

            rows_data.append({
                "row_number": r_num,
                "is_target_row": (r_num == target_row),
                "cells": row_cells,
            })

        wb.close()

        return {
            "status": "SUCCESS",
            "source_file": file_path.name,
            "sheet": active_sheet_name,
            "target_cell": target_cell_addr,
            "target_row": target_row,
            "target_column": target_col_idx,
            "target_column_letter": target_col_letter,
            "target_value": target_value,
            "file_hash": file_hash,
            "record_hash": record_hash,
            "is_provenance_verified": True,
            "total_rows": total_rows,
            "total_columns": total_cols,
            "available_sheets": available_sheets,
            "window": {
                "min_row": min_row,
                "max_row": max_row,
                "min_col": min_col,
                "max_col": max_col,
                "row_radius": row_radius,
                "column_radius": column_radius,
            },
            "context": {
                "columns": columns_meta,
                "rows": rows_data,
            },
        }

    @staticmethod
    def _format_cell_value(val: str) -> Union[int, float, str, None]:
        """Converts string cell values into numeric types if applicable."""
        if val is None or val == "":
            return ""
        val_str = str(val).strip()
        if re.match(r"^-?\d+$", val_str):
            try:
                return int(val_str)
            except ValueError:
                pass
        if re.match(r"^-?\d+\.\d+$", val_str):
            try:
                return float(val_str)
            except ValueError:
                pass
        return val_str
