import csv
import hashlib
import json
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional
import openpyxl
from pydantic import BaseModel, ConfigDict

from neofinesse.ingestion.registry import FileRegistryEntry
from neofinesse.models.base import Provider, ProvenanceReference, SourceType


def col_index_to_letter(col_idx: int) -> str:
    """Converts a 0-indexed column integer to an Excel column letter (0 -> 'A', 27 -> 'AB')."""
    result = ""
    col_idx += 1
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx - 1, 26)
        result = chr(65 + remainder) + result
    return result


class RawRecord(BaseModel):
    model_config = ConfigDict(arbitrary_types_allowed=True)

    data: Dict[str, Any]
    provenance: ProvenanceReference


class FileParser:
    """Parses CSV and Excel files while computing record hashes and exact cell coordinates."""

    def __init__(self, batch_id: str = "INGEST-BATCH-001"):
        self.batch_id = batch_id

    def compute_record_hash(self, row_data: Dict[str, Any]) -> str:
        """Computes SHA-256 hash of a single record's dictionary content."""
        serialized = json.dumps(row_data, sort_keys=True, default=str)
        return hashlib.sha256(serialized.encode("utf-8")).hexdigest()

    def parse_csv(self, entry: FileRegistryEntry) -> List[RawRecord]:
        """Parses a CSV file and constructs RawRecord objects with exact row/cell provenance."""
        filepath = Path(entry.file_path)
        records: List[RawRecord] = []

        with open(filepath, "r", encoding="utf-8") as f:
            reader = csv.reader(f)
            headers = next(reader, None)
            if not headers:
                return records

            for row_idx, row in enumerate(reader, start=2):  # Header is row 1, data starts at row 2
                row_dict = {}
                col_mapping = {}

                for col_idx, header in enumerate(headers):
                    val = row[col_idx] if col_idx < len(row) else ""
                    row_dict[header] = val
                    col_letter = col_index_to_letter(col_idx)
                    col_mapping[header] = f"{col_letter}{row_idx}"

                rec_hash = self.compute_record_hash(row_dict)

                prov = ProvenanceReference(
                    source_id=entry.source_id,
                    source_type=SourceType.CSV,
                    source_file=entry.filename,
                    source_sheet=None,
                    source_row=row_idx,
                    source_columns=col_mapping,
                    source_hash=entry.file_hash,
                    record_hash=rec_hash,
                    provider=entry.provider,
                    ingested_at=datetime.now(),
                    ingested_by=self.batch_id,
                )

                records.append(RawRecord(data=row_dict, provenance=prov))

        return records

    def parse_xlsx(self, entry: FileRegistryEntry, sheet_name: Optional[str] = None) -> List[RawRecord]:
        """Parses an XLSX Excel file preserving sheet name, row numbers, and cell coordinates."""
        filepath = Path(entry.file_path)
        wb = openpyxl.load_workbook(filepath, data_only=True)
        sheets = [sheet_name] if sheet_name else wb.sheetnames
        records: List[RawRecord] = []

        for s_name in sheets:
            ws = wb[s_name]
            rows_iter = ws.iter_rows(values_only=True)
            headers = next(rows_iter, None)
            if not headers:
                continue

            # Strip whitespace and stringify headers
            header_strs = [str(h).strip() if h is not None else f"col_{idx}" for idx, h in enumerate(headers)]

            for row_idx, row in enumerate(rows_iter, start=2):
                if not any(row):  # Skip completely empty rows
                    continue

                row_dict = {}
                col_mapping = {}

                for col_idx, header in enumerate(header_strs):
                    val = row[col_idx] if col_idx < len(row) else None
                    row_dict[header] = val
                    col_letter = col_index_to_letter(col_idx)
                    col_mapping[header] = f"{col_letter}{row_idx}"

                rec_hash = self.compute_record_hash(row_dict)

                prov = ProvenanceReference(
                    source_id=f"{entry.source_id}_{s_name.upper()}",
                    source_type=SourceType.XLSX,
                    source_file=entry.filename,
                    source_sheet=s_name,
                    source_row=row_idx,
                    source_columns=col_mapping,
                    source_hash=entry.file_hash,
                    record_hash=rec_hash,
                    provider=entry.provider,
                    ingested_at=datetime.now(),
                    ingested_by=self.batch_id,
                )

                records.append(RawRecord(data=row_dict, provenance=prov))

        return records
