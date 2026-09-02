import csv
import hashlib
import json
import os
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List
import openpyxl
from openpyxl.styles import Font, PatternFill

from neofinesse.generator.synthetic import SyntheticWorld
from neofinesse.generator.config import GeneratorConfig
from neofinesse.models.base import Provider, SourceType


class DataExporter:
    """Exports a SyntheticWorld to multi-source CSV, Excel, Registry, and isolated Ground Truth files."""

    def __init__(self, world: SyntheticWorld, config: GeneratorConfig):
        self.world = world
        self.config = config
        self.data_dir = Path(config.output_dir)
        self.gt_dir = Path(config.ground_truth_dir)

    def export_all(self) -> Dict[str, Any]:
        """Exports all datasets and returns the source registry dictionary."""
        self.data_dir.mkdir(parents=True, exist_ok=True)
        self.gt_dir.mkdir(parents=True, exist_ok=True)

        registry_entries = []

        # 1. Export CSVs
        csv_exports = [
            ("orders.csv", self._export_orders()),
            ("payments.csv", self._export_payments()),
            ("upi_transactions.csv", self._export_upi_transactions()),
            ("upi_events.csv", self._export_upi_events()),
            ("refunds.csv", self._export_refunds()),
            ("disputes.csv", self._export_disputes()),
            ("adjustments.csv", self._export_adjustments()),
            ("settlement_lines.csv", self._export_settlement_lines()),
            ("settlements.csv", self._export_settlements()),
            ("bank_transactions.csv", self._export_bank_transactions()),
        ]

        for filename, rows in csv_exports:
            filepath = self.data_dir / filename
            self._write_csv(filepath, rows)
            reg = self._create_registry_entry(filepath, SourceType.CSV, len(rows) - 1)
            registry_entries.append(reg)

        # 2. Export Excel files
        recon_xlsx = self.data_dir / "settlement_recon.xlsx"
        self._export_settlement_recon_xlsx(recon_xlsx)
        reg_recon = self._create_registry_entry(recon_xlsx, SourceType.XLSX, len(self.world.settlement_lines))
        registry_entries.append(reg_recon)

        bank_xlsx = self.data_dir / "bank_statement.xlsx"
        self._export_bank_statement_xlsx(bank_xlsx)
        reg_bank = self._create_registry_entry(bank_xlsx, SourceType.XLSX, len(self.world.bank_transactions))
        registry_entries.append(reg_bank)

        # 3. Write Source Registry JSON
        registry_path = self.data_dir / "source_registry.json"
        with open(registry_path, "w", encoding="utf-8") as f:
            json.dump(registry_entries, f, indent=2, default=str)

        # 4. Write Ground Truth JSON (in separate directory)
        gt_path = self.gt_dir / "ground_truth.json"
        gt_data = [gt.model_dump() for gt in self.world.ground_truths]
        with open(gt_path, "w", encoding="utf-8") as f:
            json.dump(gt_data, f, indent=2, default=str)

        return {
            "registry": registry_entries,
            "ground_truth_path": str(gt_path),
            "data_dir": str(self.data_dir),
        }

    def _write_csv(self, filepath: Path, rows: List[List[Any]]) -> None:
        with open(filepath, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerows(rows)

    def _create_registry_entry(self, filepath: Path, source_type: SourceType, record_count: int) -> Dict[str, Any]:
        hasher = hashlib.sha256()
        with open(filepath, "rb") as f:
            while chunk := f.read(65536):
                hasher.update(chunk)
        file_hash = hasher.hexdigest()
        file_size = filepath.stat().st_size

        return {
            "source_id": f"SRC-{filepath.stem.upper()}",
            "filename": filepath.name,
            "file_path": str(filepath),
            "file_hash": file_hash,
            "file_size": file_size,
            "format": source_type.value,
            "provider": self.config.provider.value,
            "record_count": record_count,
            "generated_at": datetime.now().isoformat(),
            "generation_seed": self.config.seed,
            "ingestion_status": "READY",
        }

    def _export_orders(self) -> List[List[Any]]:
        header = ["id", "amount", "currency", "status", "created_at"]
        rows = [header]
        for o in self.world.orders:
            rows.append([o.id, o.amount, o.currency.value, o.status, o.created_at.isoformat()])
        return rows

    def _export_payments(self) -> List[List[Any]]:
        header = [
            "id", "amount", "currency", "status", "normalized_status", "order_id",
            "method", "bank", "vpa", "fee", "tax", "net_amount", "error_code",
            "error_description", "created_at", "captured_at", "settled", "settlement_id", "provider"
        ]
        rows = [header]
        for p in self.world.payments:
            rows.append([
                p.id, p.amount, p.currency.value, p.status, p.normalized_status.value,
                p.order_id or "", p.method, p.bank or "", p.vpa or "", p.fee, p.tax, p.net_amount,
                p.error_code or "", p.error_description or "", p.created_at.isoformat(),
                p.captured_at.isoformat() if p.captured_at else "", str(p.settled).lower(),
                p.settlement_id or "", p.provider.value
            ])
        return rows

    def _export_upi_transactions(self) -> List[List[Any]]:
        header = [
            "upi_transaction_id", "payment_id", "order_id", "rrn", "amount", "vpa",
            "initiated_at", "current_observed_status", "final_determined_status",
            "debit_observed", "reversal_status", "reversal_amount", "financial_effect_status",
            "financial_effect_amount", "provider"
        ]
        rows = [header]
        for u in self.world.upi_transactions:
            rows.append([
                u.upi_transaction_id, u.payment_id, u.order_id or "", u.rrn or "", u.amount,
                u.vpa or "", u.initiated_at.isoformat(), u.current_observed_status.value,
                u.final_determined_status.value, str(u.debit_observed).lower(),
                u.reversal_status.value, u.reversal_amount or "",
                u.financial_effect_status.value, u.financial_effect_amount if u.financial_effect_amount is not None else "",
                u.provider.value
            ])
        return rows

    def _export_upi_events(self) -> List[List[Any]]:
        header = [
            "event_id", "upi_transaction_id", "timestamp", "previous_state", "new_state",
            "event_type", "amount", "rrn", "source"
        ]
        rows = [header]
        for e in self.world.upi_events:
            rows.append([
                e.event_id, e.upi_transaction_id, e.timestamp.isoformat(),
                e.previous_state.value, e.new_state.value, e.event_type,
                e.amount if e.amount is not None else "", e.rrn or "", e.source
            ])
        return rows

    def _export_refunds(self) -> List[List[Any]]:
        header = [
            "id", "amount", "currency", "payment_id", "status", "speed_requested",
            "speed_processed", "arn", "created_at", "processed_at", "settlement_id", "provider"
        ]
        rows = [header]
        for r in self.world.refunds:
            arn = r.acquirer_data.get("arn") if r.acquirer_data else ""
            rows.append([
                r.id, r.amount, r.currency.value, r.payment_id, r.status.value,
                r.speed_requested.value, r.speed_processed.value, arn or "",
                r.created_at.isoformat(), r.processed_at.isoformat() if r.processed_at else "",
                r.settlement_id or "", r.provider.value
            ])
        return rows

    def _export_disputes(self) -> List[List[Any]]:
        header = [
            "id", "payment_id", "amount", "amount_deducted", "currency", "reason_code",
            "status", "phase", "created_at", "settlement_id", "reversal_settlement_id",
            "net_financial_effect", "provider"
        ]
        rows = [header]
        for d in self.world.disputes:
            rows.append([
                d.id, d.payment_id, d.amount, d.amount_deducted, d.currency.value,
                d.reason_code or "", d.status.value, d.phase.value,
                d.created_at.isoformat(), d.settlement_id or "", d.reversal_settlement_id or "",
                d.net_financial_effect, d.provider.value
            ])
        return rows

    def _export_adjustments(self) -> List[List[Any]]:
        header = ["id", "amount", "currency", "description", "settlement_id", "adjustment_type", "created_at", "provider"]
        rows = [header]
        for a in self.world.adjustments:
            rows.append([
                a.id, a.amount, a.currency.value, a.description or "", a.settlement_id or "",
                a.adjustment_type.value, a.created_at.isoformat(), a.provider.value
            ])
        return rows

    def _export_settlement_lines(self) -> List[List[Any]]:
        header = [
            "settlement_line_id", "settlement_id", "source_event_id", "source_event_type",
            "payment_id", "amount", "fee", "tax", "net_amount", "currency",
            "event_timestamp", "settlement_timestamp", "provider"
        ]
        rows = [header]
        for l in self.world.settlement_lines:
            rows.append([
                l.settlement_line_id, l.settlement_id, l.source_event_id, l.source_event_type.value,
                l.payment_id or "", l.amount, l.fee, l.tax, l.net_amount, l.currency.value,
                l.event_timestamp.isoformat() if l.event_timestamp else "",
                l.settlement_timestamp.isoformat() if l.settlement_timestamp else "",
                l.provider.value
            ])
        return rows

    def _export_settlements(self) -> List[List[Any]]:
        header = [
            "id", "amount", "status", "fees", "tax", "utr", "gross_amount", "refund_total",
            "adjustment_total", "dispute_total", "transfer_total", "expected_amount",
            "variance", "recon_status", "created_at", "settled_at", "provider"
        ]
        rows = [header]
        for s in self.world.settlements:
            rows.append([
                s.id, s.amount, s.status.value, s.fees, s.tax, s.utr or "", s.gross_amount,
                s.refund_total, s.adjustment_total, s.dispute_total, s.transfer_total,
                s.expected_amount, s.variance, s.recon_status.value,
                s.created_at.isoformat(), s.settled_at.isoformat() if s.settled_at else "",
                s.provider.value
            ])
        return rows

    def _export_bank_transactions(self) -> List[List[Any]]:
        header = [
            "bank_txn_id", "utr", "credit_amount", "debit_amount", "balance",
            "value_date", "transaction_date", "raw_description", "parsed_utr", "account_number"
        ]
        rows = [header]
        for b in self.world.bank_transactions:
            rows.append([
                b.bank_txn_id, b.utr or "", b.credit_amount if b.credit_amount is not None else "",
                b.debit_amount if b.debit_amount is not None else "", b.balance if b.balance is not None else "",
                b.value_date.isoformat(), b.transaction_date.isoformat(),
                b.raw_description, b.parsed_utr or "", b.account_number
            ])
        return rows

    def _export_settlement_recon_xlsx(self, filepath: Path) -> None:
        wb = openpyxl.Workbook()
        # Sheet 1: Settlements
        ws_setl = wb.active
        ws_setl.title = "Settlements"
        setl_headers = ["Settlement ID", "Amount (Paise)", "Status", "Fees", "Tax", "UTR", "Expected Amount", "Variance", "Settled At"]
        ws_setl.append(setl_headers)
        for s in self.world.settlements:
            ws_setl.append([
                s.id, s.amount, s.status.value, s.fees, s.tax, s.utr or "",
                s.expected_amount, s.variance, s.settled_at.isoformat() if s.settled_at else ""
            ])

        # Sheet 2: Settlement_Lines
        ws_lines = wb.create_sheet(title="Settlement_Lines")
        line_headers = ["Line ID", "Settlement ID", "Event ID", "Event Type", "Payment ID", "Gross Amount", "Fee", "Tax", "Net Amount", "Timestamp"]
        ws_lines.append(line_headers)
        for l in self.world.settlement_lines:
            ws_lines.append([
                l.settlement_line_id, l.settlement_id, l.source_event_id, l.source_event_type.value,
                l.payment_id or "", l.amount, l.fee, l.tax, l.net_amount,
                l.event_timestamp.isoformat() if l.event_timestamp else ""
            ])

        wb.save(filepath)

    def _export_bank_statement_xlsx(self, filepath: Path) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Account_Statement"
        headers = ["Txn ID", "Value Date", "Description", "UTR", "Credit (Paise)", "Debit (Paise)", "Account Number"]
        ws.append(headers)
        for b in self.world.bank_transactions:
            ws.append([
                b.bank_txn_id, b.value_date.isoformat(), b.raw_description, b.utr or "",
                b.credit_amount if b.credit_amount is not None else 0,
                b.debit_amount if b.debit_amount is not None else 0,
                b.account_number
            ])
        wb.save(filepath)
